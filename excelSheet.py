#from xlwt import Workbook
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl


def createSheet(data): #creates the sheet using predetermined variables and settings (easy sheet)

    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.column_dimensions['A'].width=40
    for i in range(66, 74):
        col = chr(i)
        sheet1.column_dimensions[col].width = 16
    wb.save('Example Sheet.xlsx')
    colLabel(sheet1)
    fillInfo(sheet1, data)
    wb.save('Example Sheet.xlsx')
    makeLine(False,0)
    makePie(False,0)
    insGraph(sheet1,'linegraph.png','L2')
    insGraph(sheet1,'piegraph.png','L30')
    wb.save('Example Sheet.xlsx')

def makeAdvSheet(data,label,lineGraph,pieGraph,lineList,pieList):
    wb = openpyxl.Workbook()
    advsheet1=wb.active
    advsheet1.column_dimensions['A'].width = 40

    for x in range(66, 66+len(label)):
        col = chr(x)
        advsheet1.column_dimensions[col].width=16

    advSheetCol(advsheet1,label)
    fillInfo(advsheet1,data)
    wb.save('Example Sheet.xlsx')

    if(lineGraph==True):
        makeLine(True,lineList)
        insGraph(advsheet1,'linegraph.png',(chr(66+len(label))+'2'))
    if(pieGraph==True):
        makePie(True,pieList)
        if(lineGraph==False):
            insGraph(advsheet1,'piegraph.png',(chr(66+len(label))+'2'))
        else:
            insGraph(advsheet1, 'piegraph.png', (chr(66 + len(label)) + '30'))

    wb.save('Example Sheet.xlsx')

def advSheetCol(sheet,label):
    ct=0
    for x in label:
        sheet.cell(row=1, column=ct+1).value = x #label(x)
        ct+=1
    return sheet



def colLabel(sheet1):
    sheet1.cell(row=1, column=1).value = 'Item Name'
    sheet1.cell(row=1, column=2).value = 'Amount'
    sheet1.cell(row=1, column=3).value = 'Average Price'
    sheet1.cell(row=1, column=4).value = 'Total Coins'
    sheet1.cell(row=1, column=5).value = 'Average Coins'
    sheet1.cell(row=1, column=6).value = 'Percent'
    sheet1.cell(row=1, column=7).value = 'NPC Price'
    sheet1.cell(row=1, column=8).value = 'NPC Profit'
    sheet1.cell(row=1, column=9).value = 'Date/Time'
    sheet1.cell(row=1, column=10).value = 'Key'

    return sheet1


def fillInfo(sheet1, data):
    r = 0;
    c = 0;

    for i in data:
        print(i)
        for x in i:
            sheet1.cell(row=r+2,column=c+1).value = x
            c+=1
        c=0;
        r+=1


def makeLine(adv,advlist):
    var = pd.read_excel('Example Sheet.xlsx')
    plt.figure(figsize=(10,10))
    x=[]
    y=[]
    if adv==False:
        x = list(var['Date/Time'])
        y = list(var['Average Price'])
        plt.title('Average Price over Time')
        plt.xlabel('Date and Time')
        plt.ylabel('Average Price')
        plt.plot(x, y, marker='*')
        plt.xticks(np.arange(0, len(x) + 1, 50))
    else:
        x=list(var[advlist[0]])
        y=list(var[advlist[1]])
        plt.title(advlist[0] + ' over '+advlist[1])
        plt.xlabel(advlist[0])
        plt.ylabel(advlist[1])
        plt.plot(x, y, marker='*')
        plt.xticks(np.arange(0, len(x) + 1, 10))
    plt.xticks(rotation=30)
    plt.tick_params(axis='x',labelsize=10)
    plt.savefig('linegraph.png',dpi = 50)
    plt.close()


def makePie(adv,advlist):
    var = pd.read_excel('Example Sheet.xlsx')
    pieList = {}
    mylabels = []
    info = []
    plt.figure(figsize=(10,10))

    if(adv!=True):
        plt.title("Amount of Items available at Average Price")
        npc_price = list(var['Average Price'])
        amount = list(var['Amount'])
        pieInfo = tuple(zip(amount,npc_price))
        for amt,price in pieInfo:
            total = pieList.get(price,0)+amt
            pieList[price] = total
        for x,y in pieList.items():
            info.append(y)
            mylabels.append(x)
    else:
        plt.title("Relation between "+advlist[0]+" and "+advlist[1])
        group1=list(var[advlist[0]])
        group2=list(var[advlist[1]])
        pieInfo=tuple(zip(group2,group1))

        for g1,g2 in pieInfo:
            total = pieList.get(g2,0)+g1
            pieList[g2]=total
        for x,y in pieList.items():
            info.append(y)
            mylabels.append(x)

    plt.pie(info, labels=mylabels, shadow = True, autopct='%.1f%%')
    plt.savefig('piegraph.png',dpi=50)
    plt.close()

def insGraph(sheet1,image2,loc):
    img = openpyxl.drawing.image.Image(image2)
    img.anchor=loc
    sheet1.add_image(img)